# inventario/views.py
print("--- LEYENDO EL ARCHIVO VIEWS.PY CORRECTO ---")
from django.shortcuts import render, redirect, get_object_or_404
from core.decorators import role_required
from datetime import datetime
import base64
import html
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.db.models import Avg
from django.http import Http404, FileResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.models import User,Group
from django.db import connection
from django.db.models import Q
from django.db import transaction 
from core.decorators import role_required 
from django.contrib.auth.forms import UserCreationForm
from django.db.models import Q, Sum
from django.contrib.auth import update_session_auth_hash
from .utils import cifrar_dato
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from django.conf import settings
from django.http import HttpResponse
from cryptography.fernet import Fernet
from reportlab.platypus import LongTable
from reportlab.lib.pagesizes import landscape
from django.utils.html import escape
import openpyxl
from .models import Estudiante, Curso, Inscripcion, Pago, Instructor, Evaluacion
from .forms import EstudianteForm, CursoForm, InstructorForm, InscripcionForm,UsuarioForm
import traceback
import os

# --- GESTIÓN ESTUDIANTES ---
@login_required
@role_required(allowed_roles=['Admin', 'Instructor'])
def gestion_estudiantes(request):
    busqueda = request.GET.get('busqueda', '').strip().lower()
    filtro_estado = request.GET.get('estado', '').strip()
    
    estudiantes = []
    
    with connection.cursor() as cursor:
        cursor.execute("EXEC SP_Listar_Estudiantes")
        columns = [col[0].lower() for col in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            cleaned_row = []
            for i, val in enumerate(row):
                col_name = columns[i]
                if isinstance(val, bytes):
                    # Decodificar correctamente si es direccion o texto binario
                    try:
                        val = val.decode('utf-8')
                    except UnicodeDecodeError:
                        val = str(val)
                cleaned_row.append(val)
            
            estudiante_dict = dict(zip(columns, cleaned_row))
            if not estudiante_dict.get('estado'):
                estudiante_dict['estado'] = 'Activo'
            estudiantes.append(estudiante_dict)

    if busqueda:
        estudiantes = [
            e for e in estudiantes 
            if busqueda in str(e.get('nombre_completo', '')).lower() 
            or busqueda in str(e.get('email', '')).lower()
            or busqueda in str(e.get('usuario', '')).lower()
            or busqueda in str(e.get('direccion', '')).lower() # Opcional: permitir buscar por dirección también
        ]

    if filtro_estado:
        estudiantes = [
            e for e in estudiantes 
            if str(e.get('estado', '')).lower() == filtro_estado.lower()
        ]

    usuarios = User.objects.filter(groups__name='Estudiante')

    # Cálculo correcto de estadísticas para las tarjetas de la vista
    total_estudiantes = len(estudiantes)
    total_activos = sum(1 for e in estudiantes if str(e.get('estado', '')).capitalize() == 'Activo')
    total_inactivos = sum(1 for e in estudiantes if str(e.get('estado', '')).capitalize() == 'Inactivo')

    return render(request, 'inventario/estudiantes.html', {
        'estudiantes': estudiantes, 
        'usuarios': usuarios,
        'total_estudiantes': total_estudiantes,
        'total_activos': total_activos,
        'total_inactivos': total_inactivos,
    })

@login_required
@role_required(allowed_roles=['Admin'])
def crear_estudiantes(request):
    if request.method == 'POST':
        username = request.POST.get('usuario')
        password = request.POST.get('contrasena')
        nombre = request.POST.get('nombre_completo', '').strip().upper()
        email = request.POST.get('email')
        estado_val = request.POST.get('estado', 'Activo')
        num_doc = request.POST.get('numero_documento', '').strip()
        tipo_doc = request.POST.get('tipo_documento', '').upper()
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion', '').upper()

        try:
            with transaction.atomic():
                # Dividir el nombre completo en partes para guardarlo en el modelo User de Django
                partes_nombre = nombre.split(' ', 1)
                primer_nombre = partes_nombre[0]
                apellido = partes_nombre[1] if len(partes_nombre) > 1 else ''

                # 1. Crear el usuario de autenticación de Django incluyendo el nombre
                user = User.objects.create_user(
                    username=username, 
                    email=email, 
                    first_name=primer_nombre, 
                    last_name=apellido
                )
                user.set_password(password)
                user.save()

                # 2. Asignar automáticamente el grupo "Estudiante" para que no aparezca "Sin rol"
                try:
                    grupo_estudiante = Group.objects.get(name='Estudiante')
                    user.groups.add(grupo_estudiante)
                except Group.DoesNotExist:
                    grupo_estudiante = Group.objects.create(name='Estudiante')
                    user.groups.add(grupo_estudiante)

                contra_cifrada = password.encode('utf-8') if password else b''

                # 3. Insertar directamente con SQL crudo utilizando EncryptByKey y usuario_id
                with connection.cursor() as cursor:
                    cursor.execute("""
                        OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;
                        
                        INSERT INTO Estudiantes (
                            usuario_id, usuario, contrasena, nombre_completo, 
                            email, telefono, direccion, tipo_documento, numero_documento, estado
                        ) 
                        VALUES (
                            %s, %s, %s, %s, 
                            %s, %s, %s, %s, EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), %s)), %s
                        );
                        
                        CLOSE SYMMETRIC KEY ClaveDatos;
                    """, [
                        user.id,        # Corresponde a usuario_id
                        username, 
                        contra_cifrada, 
                        nombre, 
                        email, 
                        telefono, 
                        direccion, 
                        tipo_doc, 
                        num_doc,  
                        estado_val
                    ])

                messages.success(request, 'Estudiante registrado correctamente.')
                
        except Exception as e:
            print("==================================================")
            print("--- ERROR CRÍTICO AL CREAR ESTUDIANTE:", str(e))
            print("==================================================")
            messages.error(request, f'Error al registrar: {e}')
            
    return redirect('inventario:gestion_estudiantes')

registrar_estudiante = crear_estudiantes
@login_required
def editar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # 1. Capturamos los valores del formulario
                estudiante.nombre_completo = request.POST.get('nombre_completo', '').strip()
                estudiante.email = request.POST.get('email', '').strip()
                estudiante.telefono = request.POST.get('telefono', '').strip()
                estudiante.tipo_documento = request.POST.get('tipo_documento', '').strip()
                estudiante.estado = request.POST.get('estado', 'Activo').strip()  # <--- ¡Aquí toma el Activo/Inactivo perfectamente!
                estudiante.direccion = request.POST.get('direccion', '').strip()
                
                # Guardamos los campos de texto comunes usando el ORM de Django
                estudiante.save(update_fields=['nombre_completo', 'email', 'telefono', 'tipo_documento', 'estado', 'direccion'])

                # 2. Si el usuario ingresó un NUEVO número de documento (que requiere cifrado simétrico)
                nuevo_num_doc = request.POST.get('numero_documento', '').strip()
                if nuevo_num_doc:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;
                            
                            UPDATE Estudiantes 
                            SET numero_documento = EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), ?))
                            WHERE estudiante_id = ?;
                            
                            CLOSE SYMMETRIC KEY ClaveDatos;
                        """, [nuevo_num_doc, estudiante_id])

                # 3. Sincronizamos con el usuario de autenticación si existe
                if hasattr(estudiante, 'usuario_auth') and estudiante.usuario_auth:
                    user = estudiante.usuario_auth
                    partes = estudiante.nombre_completo.split(' ', 1)
                    user.first_name = partes[0]
                    user.last_name = partes[1] if len(partes) > 1 else ''
                    user.email = estudiante.email
                    user.save()

                messages.success(request, 'Estudiante actualizado correctamente.')
                return redirect('inventario:gestion_estudiantes')
                
        except Exception as e:
            messages.error(request, f'Error al actualizar en la base de datos: {e}')
            
    return render(request, 'inventario/editar_estudiante.html', {
        'estudiante': estudiante
    })
@login_required
@role_required(allowed_roles=['Admin'])
def eliminar_estudiante(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, pk=estudiante_id)
    
    if request.method == 'POST':
        usuario_auth = estudiante.usuario_auth
        estudiante.delete()
        
        if usuario_auth:
            usuario_auth.delete()
            
        messages.success(request, 'Estudiante eliminado correctamente.')
        return redirect('inventario:gestion_estudiantes')
        
    return render(request, 'inventario/confirmar_eliminacion.html', {'objeto': estudiante, 'tipo': 'estudiante'})

# --- GESTIÓN CURSOS ---
@login_required
def gestion_cursos(request):
    # Definimos la consulta de instructores fuera para tenerla disponible siempre
    instructores = Instructor.objects.all()

    if request.method == 'POST':
        form = CursoForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    nuevo_curso = form.save(commit=False)
                    nuevo_curso.cupo_disponible = nuevo_curso.cupo_maximo
                    nuevo_curso.save()
                messages.success(request, 'Curso creado exitosamente.')
                return redirect('inventario:gestion_cursos')
            except Exception as e:
                messages.error(request, f'Error al registrar el curso: {e}')
        else:
            messages.error(request, 'Error al crear el curso. Revisa los datos.')
    else:
        form = CursoForm()

    # CORREGIDO: Añadido .select_related('instructor') para traer los datos del instructor junto al curso
    base_queryset = Curso.objects.select_related('instructor').all() if request.user.is_superuser else Curso.objects.select_related('instructor').filter(instructor__usuario_auth=request.user)
    
    # Capturamos los parámetros de búsqueda y filtro de la URL (GET)
    query = request.GET.get('q', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # Aplicamos los filtros sobre una copia del queryset para la tabla
    cursos_filtrados = base_queryset
    if query:
        cursos_filtrados = cursos_filtrados.filter(nombre_curso__icontains=query)
    
    if estado_filtro:
        cursos_filtrados = cursos_filtrados.filter(estado=estado_filtro)

    context = {
        'cursos': cursos_filtrados,
        'form': form,
        'instructores': instructores,
        'total_cursos': base_queryset.count(),
        'total_activos': base_queryset.filter(estado='Activo').count(),
        'costo_activos': base_queryset.filter(estado='Activo').aggregate(Sum('costo'))['costo__sum'] or 0,
        'costo_total': base_queryset.aggregate(Sum('costo'))['costo__sum'] or 0
    }
    
    return render(request, 'inventario/cursos.html', context)
  
@login_required
@role_required(allowed_roles=['Admin'])
def crear_curso(request):
    if request.method == 'POST':
        try:
            Curso.objects.create(
                nombre_curso=request.POST.get('nombre_curso'),
                categoria=request.POST.get('categoria'),
                duracion_horas=request.POST.get('duracion_horas'),
                costo=request.POST.get('costo'),
                estado=request.POST.get('estado'),
                cupo_maximo=request.POST.get('cupo_maximo'),
                instructor_id=request.POST.get('instructor')
            )
            messages.success(request, 'Curso registrado exitosamente.')
        except Exception as e:
            messages.error(request, f'Error al registrar el curso: {e}')
            
    return redirect('inventario:gestion_cursos')

@login_required
@role_required(allowed_roles=['Admin'])
def editar_curso(request, curso_id):
    # Buscamos el curso existente
    curso = get_object_or_404(Curso, pk=curso_id)
    
    if request.method == 'POST':
        form = CursoForm(request.POST, instance=curso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Curso actualizado.')
            return redirect('inventario:gestion_cursos')
        else:
            # ESTO TE PERMITIRÁ VER EL ERROR EXACTO EN LA TERMINAL DE DJANGO
            print("Errores en el formulario de curso:", form.errors)
            messages.error(
                request, 'Por favor, corrige los errores en el formulario.'
            )
    else:
        form = CursoForm(instance=curso)
    return render(request, 'inventario/editar_curso.html', {'form': form, 'curso': curso})

@login_required
@role_required(allowed_roles=['Admin'])
def eliminar_curso(request, curso_id):
    curso = get_object_or_404(Curso, pk=curso_id)
    if request.method == 'POST':
        curso.delete()
        messages.success(request, 'Curso eliminado correctamente.')
        return redirect('inventario:gestion_cursos')
    return render(request, 'inventario/confirmar_eliminacion.html', {'objeto': curso, 'tipo': 'curso'})


# --- GESTIÓN PAGOS ---
@login_required
@role_required(allowed_roles=['Admin'])
def gestion_pagos(request):
    # 1. Traemos los pagos optimizando las relaciones para evitar que el Estudiante salga como N/A
    pagos = Pago.objects.select_related(
        'inscripcion', 
        'inscripcion__estudiante', 
        'inscripcion__curso'
    ).all()

    # 2. Capturamos los filtros enviados por el formulario (GET)
    curso_id = request.GET.get('curso')
    estado = request.GET.get('estado')
    query = request.GET.get('q')

    # 3. Aplicamos los filtros condicionalmente si el usuario los seleccionó/escribió
    if curso_id:
        pagos = pagos.filter(inscripcion__curso_id=curso_id)
        
    if estado:
        pagos = pagos.filter(estado=estado)
        
    if query:
        # Busca por folio de inscripción, nombre completo del estudiante o texto libre
        pagos = pagos.filter(
            Q(inscripcion__folio_inscripcion__icontains=query) | 
            Q(inscripcion__estudiante__nombre_completo__icontains=query) |
            Q(metodo_pago__icontains=query)
        )

    # 4. Datos adicionales para las tarjetas de resumen y los selects del HTML
    total_pagos = pagos.count()
    pagos_registrados = pagos.filter(estado='Pagado').count()
    pagos_cancelados = pagos.filter(estado='Cancelada').count() 
    cursos = Curso.objects.all().order_by('nombre_curso')
    
    # Obtenemos los estados disponibles directamente del modelo para mantener consistencia
    estado_choices = [
        ('Pagado', 'Pagado'),
        ('Cancelada', 'Cancelada'),
    ]

    context = {
        'pagos': pagos,
        'cursos': cursos,
        'estado_choices': estado_choices,
        'total_pagos': total_pagos,
        'pagos_registrados': pagos_registrados,
        'pagos_cancelados': pagos_cancelados,
    }

    return render(request, 'inventario/pagos.html', context)
@login_required
@role_required(allowed_roles=['Admin']) # O ['Administrador'] según corresponda en tu BD
def admin_descargar_comprobante_pdf(request, pago_id):
    # 1. Obtener datos del pago (sin descifrar para obtener los bytes cifrados)
    with connection.cursor() as cursor:
        try:
            query = f"""
                SELECT 
                    p.pago_id,
                    c.nombre_curso,
                    p.fecha_pago,
                    p.monto,
                    p.estado,
                    p.referencia_pago,
                    e.nombre_completo
                FROM Pagos p
                INNER JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id
                INNER JOIN Cursos c ON i.curso_id = c.curso_id
                INNER JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
                WHERE p.pago_id = {int(pago_id)}
            """
            cursor.execute(query)
            row = cursor.fetchone()
        finally:
            pass

    if not row:
        raise Http404("El comprobante no existe.")

    # Procesar los bytes cifrados de la referencia para mostrarlos como cadena de bytes en el PDF
    ref_raw = row[5]
    referencia_cifrada = "N/A"
    
    if ref_raw:
        try:
            if isinstance(ref_raw, memoryview):
                referencia_cifrada = str(ref_raw.tobytes())
            elif isinstance(ref_raw, bytes):
                referencia_cifrada = str(ref_raw)
            else:
                referencia_cifrada = str(ref_raw)
        except Exception:
            referencia_cifrada = "Error al procesar referencia"

    pago_data = {
        'pago_id': row[0],
        'curso': row[1],
        'fecha': row[2],
        'monto': row[3],
        'estado': row[4],
        'referencia': referencia_cifrada,
        'estudiante': row[6]
    }

    # 2. Definir la ruta de almacenamiento en media/reports
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"comprobante_pago_admin_{pago_data['pago_id']}.pdf"
    file_path = os.path.join(reports_dir, filename)

    # 3. Generación del documento PDF con ReportLab
    c = canvas.Canvas(file_path, pagesize=letter)
    width, height = letter

    # Encabezado estético
    c.setFillColor(colors.HexColor("#312e81"))
    c.rect(0, height - 100, width, 100, fill=1, stroke=0)

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(50, height - 45, "SIPEL - COMPROBANTE DE PAGO")
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 70, f"Comprobante Oficial N°: 0000{pago_data['pago_id']}")

    # Cuerpo / Información del pago
    c.setFillColor(colors.HexColor("#1f2937"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 140, "Detalles de la Transacción:")

    c.setFont("Helvetica", 12)
    y_position = height - 180
    space = 28

    detalles = [
        ("Estudiante:", pago_data['estudiante']),
        ("Curso Adquirido:", pago_data['curso']),
        ("Fecha y Hora:", pago_data['fecha'].strftime("%d/%m/%Y %H:%M") if pago_data['fecha'] else "N/A"),
        ("Monto Cancelado:", f"${pago_data['monto']}"),
        ("Estado del Pago:", pago_data['estado'].upper() if pago_data['estado'] else "N/A"),
        ("Referencia de Pago:", pago_data['referencia'])
    ]

    for label, value in detalles:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y_position, label)
        
        # Fuente más pequeña para que el texto cifrado largo encaje mejor en el PDF del admin
        c.setFont("Helvetica", 9)
        c.drawString(200, y_position, str(value))
        y_position -= space

    # Línea divisoria inferior
    c.setStrokeColor(colors.HexColor("#d1d5db"))
    c.setLineWidth(1)
    c.line(50, 130, width - 50, 130)

    # Pie de página
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(colors.HexColor("#6b7280"))
    c.drawString(50, 100, "Este documento es un comprobante digital generado automáticamente por el sistema SIPEL.")
    c.drawString(50, 85, "Conserve este archivo para cualquier aclaración futura.")

    c.showPage()
    c.save()

    # 4. Servir el archivo generado
    return FileResponse(open(file_path, 'rb'), as_attachment=False, filename=filename)
#instructores-------------------------
@login_required
@role_required(allowed_roles=['Admin'])
def gestion_instructores(request):
    query = request.GET.get('q', '')
    especialidad_filtro = request.GET.get('especialidad', '')
    estado_filtro = request.GET.get('estado', '')

    instructores = Instructor.objects.all()

    if query:
        instructores = instructores.filter(
            Q(nombre_completo__icontains=query) | 
            Q(email__icontains=query) | 
            Q(especialidad__icontains=query)
        )
    
    if especialidad_filtro:
        instructores = instructores.filter(especialidad=especialidad_filtro)
        
    if estado_filtro:
        instructores = instructores.filter(estado=estado_filtro)

    total_instructores = Instructor.objects.count()
    total_activos = Instructor.objects.filter(estado='Activo').count()
    total_inactivos = Instructor.objects.filter(estado='Inactivo').count()

    lista_especialidades = Instructor.objects.values_list('especialidad', flat=True).distinct()

    context = {
        'instructores': instructores,
        'total_instructores': total_instructores,
        'total_activos': total_activos,
        'total_inactivos': total_inactivos,
        'lista_especialidades': lista_especialidades,
    }
    
    return render(request, 'inventario/instructores.html', context)
@login_required
@role_required(allowed_roles=['Admin'])
def crear_instructor(request):
    if request.method == 'POST':
        username = request.POST.get('usuario')
        password = request.POST.get('contrasena')
        nombre = request.POST.get('nombre_completo', '').strip().upper()
        email = request.POST.get('email')
        estado_val = request.POST.get('estado', 'Activo')
        cedula_raw = request.POST.get('cedula_profesional', '').strip()
        especialidad = request.POST.get('especialidad', '').upper().strip()
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion', '').upper().strip()

        try:
            with transaction.atomic():
                # Dividir el nombre completo en partes para guardarlo en el modelo User de Django
                partes_nombre = nombre.split(' ', 1)
                primer_nombre = partes_nombre[0]
                apellido = partes_nombre[1] if len(partes_nombre) > 1 else ''

                # 1. Crear el usuario de autenticación de Django incluyendo el nombre
                user = User.objects.create_user(
                    username=username, 
                    email=email, 
                    first_name=primer_nombre, 
                    last_name=apellido
                )
                user.set_password(password)
                user.save()

                # 2. Asignar automáticamente el grupo "Instructor" para que no aparezca "Sin rol"
                try:
                    grupo_instructor = Group.objects.get(name='Instructor')
                    user.groups.add(grupo_instructor)
                except Group.DoesNotExist:
                    grupo_instructor = Group.objects.create(name='Instructor')
                    user.groups.add(grupo_instructor)

                contra_cifrada = password.encode('utf-8') if password else b''

                # 3. Insertar directamente con SQL crudo utilizando EncryptByKey y usuario_id
                with connection.cursor() as cursor:
                    cursor.execute("""
                        OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;
                        
                        INSERT INTO Instructores (
                            usuario_id, usuario, contrasena, nombre_completo, 
                            email, telefono, direccion, especialidad, cedula_profesional, estado
                        ) 
                        VALUES (
                            %s, %s, %s, %s, 
                            %s, %s, %s, %s, EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), %s)), %s
                        );
                        
                        CLOSE SYMMETRIC KEY ClaveDatos;
                    """, [
                        user.id,        # Corresponde a usuario_id
                        username, 
                        contra_cifrada, 
                        nombre, 
                        email, 
                        telefono, 
                        direccion, 
                        especialidad, 
                        cedula_raw,  
                        estado_val
                    ])

                messages.success(request, 'Instructor registrado correctamente.')
                
        except Exception as e:
            print("==================================================")
            print("--- ERROR CRÍTICO AL CREAR INSTRUCTOR:", str(e))
            print("==================================================")
            messages.error(request, f'Error al registrar: {e}')
            
    return redirect('inventario:gestion_instructores')

registrar_instructor = crear_instructor
@login_required
@role_required(allowed_roles=['Admin'])
def editar_instructor(request, instructor_id):
    instructor = get_object_or_404(Instructor, pk=instructor_id)
    
    if request.method == 'POST':
        form = InstructorForm(request.POST, instance=instructor)
        
        try:
            with transaction.atomic():
                # 1. Extraemos los valores directamente del diccionario request.POST
                nombre = request.POST.get('nombre_completo', '').strip().upper()
                email = request.POST.get('email', '').strip()
                telefono = request.POST.get('telefono', '').strip()
                especialidad = request.POST.get('especialidad', '').upper().strip()
                estado = request.POST.get('estado', 'Activo').strip()
                direccion = request.POST.get('direccion', '').upper().strip()
                usuario = request.POST.get('usuario', '').strip()
                nueva_cedula = request.POST.get('cedula_profesional', '').strip()
                
                # Capturamos la contraseña enviada en el formulario
                nueva_password = request.POST.get('password', '').strip()
                
                with connection.cursor() as cursor:
                    # 2. Si el usuario ingresó una nueva cédula, la ciframos con SQL Server
                    if nueva_cedula:
                        cursor.execute("""
                            OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;
                            
                            UPDATE Instructores 
                            SET nombre_completo = %s,
                                email = %s,
                                telefono = %s,
                                direccion = %s,
                                usuario = %s,
                                especialidad = %s,
                                estado = %s,
                                cedula_profesional = EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), %s))
                            WHERE instructor_id = %s;
                            
                            CLOSE SYMMETRIC KEY ClaveDatos;
                        """, [nombre, email, telefono, direccion, usuario, especialidad, estado, nueva_cedula, instructor_id])
                    
                    # 3. Si dejó la cédula en blanco, actualizamos el resto de campos sin tocar la cédula actual
                    else:
                        cursor.execute("""
                            UPDATE Instructores 
                            SET nombre_completo = %s,
                                email = %s,
                                telefono = %s,
                                direccion = %s,
                                usuario = %s,
                                especialidad = %s,
                                estado = %s
                            WHERE instructor_id = %s;
                        """, [nombre, email, telefono, direccion, usuario, especialidad, estado, instructor_id])

                # 4. Sincronizamos con el usuario de autenticación (auth_user)
                if hasattr(instructor, 'usuario_auth') and instructor.usuario_auth:
                    user = instructor.usuario_auth
                    partes = nombre.strip().split(' ', 1)
                    user.first_name = partes[0]
                    user.last_name = partes[1] if len(partes) > 1 else ''
                    user.email = email
                    
                    # Actualizamos el username si cambió
                    if usuario:
                        user.username = usuario
                        
                    # AQUÍ ESTABA LA CLAVE: Actualizamos la contraseña de forma segura para que Django la reconozca en el login
                    if nueva_password:
                        user.set_password(nueva_password)
                        
                    user.save()

            messages.success(request, 'Instructor actualizado correctamente.')
            return redirect('inventario:gestion_instructores')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar en la base de datos: {e}')
    else:
        form = InstructorForm(instance=instructor)
        if 'cedula_profesional' in form.fields:
            form.fields['cedula_profesional'].initial = ''
        
    return render(request, 'inventario/editar_instructor.html', {
        'form': form,
        'instructor': instructor
    })

@login_required
@role_required(allowed_roles=['Admin'])
def eliminar_instructor(request, instructor_id):
    instructor = get_object_or_404(Instructor, pk=instructor_id)
    
    if request.method == 'POST':
        usuario_auth = getattr(instructor, 'usuario_auth', None)
        instructor.delete()
        
        if usuario_auth:
            usuario_auth.delete()
            
        messages.success(request, 'Instructor eliminado correctamente.')
        return redirect('inventario:gestion_instructores')
        
    return render(request, 'inventario/confirmar_eliminacion.html', {'objeto': instructor, 'tipo': 'instructor'})

# --- INSCRIPCIONES ---
@login_required
@role_required(allowed_roles=['Admin', 'Estudiante'])
def gestion_inscripciones(request):
    # Partimos de todas las inscripciones con las relaciones optimizadas
    inscripciones = Inscripcion.objects.select_related('estudiante', 'curso', 'instructor').all()
    
    # Capturamos los parámetros de búsqueda del método GET
    query_estudiante = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    estado_filtro = request.GET.get('estado', '').strip()

    # Aplicamos filtro por estudiante (buscando por nombre completo)
    if query_estudiante:
        inscripciones = inscripciones.filter(
            Q(estudiante__nombre_completo__icontains=query_estudiante)
        )

    # Aplicamos filtro por curso
    if curso_id:
        inscripciones = inscripciones.filter(curso__curso_id=curso_id)

    # Aplicamos filtro por estado ('Pagado' o 'Cancelada')
    if estado_filtro:
        inscripciones = inscripciones.filter(estado=estado_filtro)

    # Listas para poblar los selectores del formulario
    estudiantes = Estudiante.objects.all()
    cursos = Curso.objects.all()
    instructores = Instructor.objects.all()
    
    # Cálculo de estadísticas generales (basado en toda la tabla sin filtros aplicados)
    total_inscripciones = Inscripcion.objects.count()
    total_pagados = Inscripcion.objects.filter(estado='Pagado').count()
    total_canceladas = Inscripcion.objects.filter(estado='Cancelada').count()

    return render(request, 'inventario/inscripciones.html', {
        'inscripciones': inscripciones,
        'estudiantes': estudiantes,
        'cursos': cursos,
        'instructores': instructores,
        'total_inscripciones': total_inscripciones,
        'total_pagados': total_pagados,
        'total_canceladas': total_canceladas,
    })
@login_required
@role_required(allowed_roles=['Admin'])
def crear_inscripciones(request):
    if request.method == 'POST':
        try:
            estudiante_id = request.POST.get('estudiante')
            curso_id = request.POST.get('curso')
            instructor_id = request.POST.get('instructor')
            estado = request.POST.get('estado') # Capturará 'Pagado' o 'Cancelado'
            
            metodo_pago = request.POST.get('metodo_pago')
            referencia_pago = request.POST.get('folio_inscripcion', '').strip()
            
            # Si la referencia está vacía, se puede asignar un valor por defecto o dejarla limpia
            if not referencia_pago:
                referencia_pago = "EFECTIVO"

            curso_actual = get_object_or_404(Curso, pk=curso_id)
            
            import random
            folio_generado = random.randint(2000, 9999)

            estudiante_db_id = int(estudiante_id)
            curso_db_id = int(curso_actual.curso_id)
            instructor_db_id = int(instructor_id) if instructor_id else "NULL"
            folio_db = int(folio_generado)
            costo_db = float(curso_actual.costo) if hasattr(curso_actual, 'costo') else 0.00
            
            metodo_clean = metodo_pago.replace("'", "''") if metodo_pago else ""
            referencia_clean = referencia_pago.replace("'", "''") if referencia_pago else ""
            estado_clean = estado.replace("'", "''") if estado else "Pagado"

            sql_ejecucion = f"""
                EXEC sp_RegistrarInscripcionConPago 
                    @estudiante_id = {estudiante_db_id}, 
                    @curso_id = {curso_db_id}, 
                    @instructor_id = {instructor_db_id}, 
                    @folio_inscripcion = {folio_db}, 
                    @total_pago = {costo_db}, 
                    @estado_inscripcion = '{estado_clean}', 
                    @metodo_pago = '{metodo_clean}', 
                    @referencia_pago = '{referencia_clean}';
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_ejecucion)
            
            if estado == 'Cancelado':
                messages.warning(request, 'La inscripción del estudiante ha sido marcada como cancelada.')
            else:
                messages.success(request, 'Inscripción y pago registrados correctamente por el administrador.')
            
            return redirect('inventario:gestion_inscripciones')
                
        except Exception as e:
            print("--- ERROR DETALLADO EN INSCRIPCIÓN ADMIN ---")
            import traceback
            traceback.print_exc()
            from django.http import HttpResponse
            return HttpResponse(f"<h1>Error atrapado:</h1><pre>{e}</pre>", status=500)
            
    return redirect('inventario:gestion_inscripciones')

@login_required
@role_required(allowed_roles=['Admin'])
def editar_inscripcion(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, pk=inscripcion_id)
    if request.method == 'POST':
        form = InscripcionForm(request.POST, instance=inscripcion)
        if form.is_valid():
            nuevo_estado = form.cleaned_data['estado']
            
            # 1. Actualizamos la inscripción mediante QuerySet.update() para evitar conflictos de ID en SQL Server
            Inscripcion.objects.filter(pk=inscripcion_id).update(
                estudiante=form.cleaned_data['estudiante'],
                curso=form.cleaned_data['curso'],
                instructor=form.cleaned_data['instructor'],
                folio_inscripcion=form.cleaned_data['folio_inscripcion'],
                estado=nuevo_estado,
                total_pago=form.cleaned_data['total_pago']
            )
            
            # 2. Sincronizamos automáticamente el estado en la tabla de Pagos relacionada
            from inventario.models import Pago
            Pago.objects.filter(inscripcion_id=inscripcion_id).update(
                estado=nuevo_estado
            )
            
            messages.success(request, 'Inscripción y su pago asociado actualizados correctamente.')
            return redirect('inventario:gestion_inscripciones')
    else:
        form = InscripcionForm(instance=inscripcion)
        
    return render(request, 'inventario/editar_inscripcion.html', {'form': form, 'inscripcion': inscripcion})



@login_required
@role_required(allowed_roles=['Admin'])
def eliminar_inscripcion(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, pk=inscripcion_id)
    if request.method == 'POST':
        inscripcion.delete()
        messages.success(request, 'Inscripción eliminada correctamente.')
        return redirect('inventario:gestion_inscripciones')
    return render(request, 'inventario/confirmar_eliminacion.html', {'objeto': inscripcion, 'tipo': 'inscripción'})


# --- EVALUACIONES Y OTROS ---
@login_required
@role_required(allowed_roles=['Admin'])
def gestion_evaluaciones(request):
    query = request.GET.get('q', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    
    evaluaciones = []
    cursos = []
    
    with connection.cursor() as cursor:
        # Obtener lista de cursos para el menú desplegable del buscador
        cursor.execute("SELECT curso_id, nombre_curso FROM Cursos")
        cursos_cols = [col[0] for col in cursor.description]
        cursos = [dict(zip(cursos_cols, row)) for row in cursor.fetchall()]
        
        # Construcción de la consulta SQL utilizando CAST a VARBINARY(MAX) para los comentarios cifrados
        sql = """
            SELECT 
                ev.evaluacion_id,
                CONCAT(au.first_name, ' ', au.last_name) AS estudiante,
                c.nombre_curso,
                ev.calificacion,
                CAST(ev.comentarios AS VARBINARY(MAX)) AS comentarios,
                ev.fecha_evaluacion
            FROM Evaluaciones ev
            INNER JOIN Inscripciones i ON ev.inscripcion_id = i.inscripcion_id
            INNER JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
            INNER JOIN auth_user au ON e.usuario_id = au.id
            INNER JOIN Cursos c ON i.curso_id = c.curso_id
            WHERE 1=1
        """
        params = []
        
        if query:
            sql += " AND (CONCAT(au.first_name, ' ', au.last_name) LIKE %s OR au.first_name LIKE %s OR au.last_name LIKE %s)"
            params.extend([f"%{query}%", f"%{query}%", f"%{query}%"])
            
        if curso_id:
            sql += " AND c.curso_id = %s"
            params.append(curso_id)
            
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            cleaned_row = []
            fila_dict = dict(zip(columns, row))
            
            for col_name, val in fila_dict.items():
                if col_name == 'comentarios':
                    if isinstance(val, memoryview):
                        val = str(val.tobytes())
                    elif isinstance(val, bytes):
                        val = str(val)
                    else:
                        val = str(val) if val is not None else "N/A"
                elif isinstance(val, bytes):
                    try:
                        val = val.decode('utf-8', errors='ignore')
                    except Exception:
                        val = str(val)
                cleaned_row.append(val)
                
            evaluaciones.append(dict(zip(columns, cleaned_row)))

    return render(request, 'inventario/evaluaciones.html', {
        'evaluaciones': evaluaciones,
        'cursos': cursos,
        'query': query,
        'curso_seleccionado': curso_id
    })
@login_required
@role_required(allowed_roles=['Admin'])
def editar_evaluacion(request, pk):
    with connection.cursor() as cursor:
        # Abrir llave para poder leer los comentarios cifrados al momento de editar
        cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
        cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
        
        if request.method == 'POST':
            calificacion = request.POST.get('calificacion')
            comentarios = request.POST.get('comentarios')
            
            # Actualizar cifrando nuevamente el comentario si es necesario o directo
            cursor.execute("""
                UPDATE Evaluaciones 
                SET calificacion = %s, comentarios = EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARCHAR(MAX), %s))
                WHERE evaluacion_id = %s
            """, [calificacion, comentarios, pk])
            
            cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")
            return redirect('inventario:gestion_evaluaciones')
            
        # Obtener los datos actuales de la evaluación para mostrarlos en el formulario
        cursor.execute("""
            SELECT evaluacion_id, calificacion, CAST(DecryptByKey(comentarios) AS VARCHAR(MAX)) AS comentarios 
            FROM Evaluaciones WHERE evaluacion_id = %s
        """, [pk])
        row = cursor.fetchone()
        
        cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")
        
    evaluacion = {
        'evaluacion_id': row[0],
        'calificacion': row[1],
        'comentarios': row[2]
    }
    
    return render(request, 'inventario/editar_evaluacion.html', {'evaluacion': evaluacion})

@login_required
@role_required(allowed_roles=['Admin'])
def editar_evaluacion(request, pk):
    with connection.cursor() as cursor:
        # Abrir llave para poder leer y/o actualizar los comentarios cifrados
        cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
        cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
        
        if request.method == 'POST':
            calificacion = request.POST.get('calificacion')
            comentarios = request.POST.get('comentarios', '').strip()
            
            if comentarios:
                # Si el administrador escribió un nuevo comentario, lo ciframos y actualizamos ambos campos
                cursor.execute("""
                    UPDATE Evaluaciones 
                    SET calificacion = %s, comentarios = EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARCHAR(MAX), %s))
                    WHERE evaluacion_id = %s
                """, [calificacion, comentarios, int(pk)])
            else:
                # Si el campo se dejó vacío, actualizamos SOLO la calificación y dejamos intacto el comentario anterior
                cursor.execute("""
                    UPDATE Evaluaciones 
                    SET calificacion = %s
                    WHERE evaluacion_id = %s
                """, [calificacion, int(pk)])
            
            cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")
            return redirect('inventario:gestion_evaluaciones')
            
        # Obtener los datos actuales de la evaluación descifrados exclusivamente para el formulario de edición
        cursor.execute("""
            SELECT evaluacion_id, calificacion, CAST(DecryptByKey(comentarios) AS VARCHAR(MAX)) AS comentarios 
            FROM Evaluaciones WHERE evaluacion_id = %s
        """, [int(pk)])
        row = cursor.fetchone()
        
        cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")
        
    evaluacion = {
        'evaluacion_id': row[0],
        'calificacion': row[1],
        'comentarios': row[2] if row[2] is not None else ""
    }
    
    return render(request, 'inventario/editar_evaluacion.html', {'evaluacion': evaluacion})
@login_required
@role_required(allowed_roles=['Admin'])
def eliminar_evaluacion(request, pk):
    if request.method == 'POST':
        with connection.cursor() as cursor:
            try:
                # Opcional pero recomendado si hay triggers o restricciones asociadas a campos cifrados
                cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
                cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
                
                # Ejecutar la eliminación del registro
                cursor.execute("DELETE FROM Evaluaciones WHERE evaluacion_id = %s", [int(pk)])
            finally:
                try:
                    cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")
                except Exception:
                    pass
                    
        return redirect('inventario:gestion_evaluaciones')
        
    return render(request, 'inventario/confirmar_eliminacion.html', {'pk': pk})

#Reportes---------------------------------

def cifrar_valor(valor):
    """Función auxiliar para cifrar datos sensibles antes de almacenarlos o procesarlos."""
    if not valor:
        return valor
    try:
        f = Fernet(settings.ENCRYPTION_KEY.encode() if isinstance(settings.ENCRYPTION_KEY, str) else settings.ENCRYPTION_KEY)
        val_bytes = str(valor).encode('utf-8')
        return f.encrypt(val_bytes).decode('utf-8')
    except Exception:
        return valor

@login_required
def reportes_view(request):
    tipo_reporte = request.GET.get('tipo_reporte', 'estudiantes')
    fecha_inicio = request.GET.get('fecha_inicio', '2026-01-01')
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    resultados = []
    columnas = []

    # Variables iniciales para las gráficas de los 6 módulos
    total_activos = 0
    total_inactivos = 0
    
    cursos_labels = []
    cursos_data = []
    
    instructores_labels = []
    instructores_data = []
    
    inscripciones_labels = []
    inscripciones_data = []
    
    eval_aprobadas = 0
    eval_reprobadas = 0
    
    pagos_labels = []
    pagos_data = []

    with connection.cursor() as cursor:
        
        # 1. Gráfica Estudiantes (Activos vs Inactivos)
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN estado = 'Activo' THEN 1 ELSE 0 END),
                SUM(CASE WHEN estado != 'Activo' THEN 1 ELSE 0 END)
            FROM Estudiantes
        """)
        row_est = cursor.fetchone()
        if row_est:
            total_activos = row_est[0] or 0
            total_inactivos = row_est[1] or 0

        # 2. Gráfica Cursos (Más demandados)
        cursor.execute("""
            SELECT TOP 5 c.nombre_curso, COUNT(i.inscripcion_id) as total_inscritos
            FROM Cursos c
            LEFT JOIN Inscripciones i ON c.curso_id = i.curso_id
            GROUP BY c.nombre_curso
            ORDER BY total_inscritos DESC
        """)
        for rc in cursor.fetchall():
            cursos_labels.append(rc[0])
            cursos_data.append(rc[1])

        # 3. Gráfica Instructores (Carga académica / Cursos asignados por instructor)
        cursor.execute("""
            SELECT TOP 5 ins.nombre_completo, COUNT(c.curso_id) as total_cursos
            FROM Instructores ins
            LEFT JOIN Cursos c ON ins.instructor_id = c.instructor_id
            GROUP BY ins.nombre_completo
            ORDER BY total_cursos DESC
        """)
        for ri in cursor.fetchall():
            instructores_labels.append(ri[0])
            instructores_data.append(ri[1])

        # 4. Gráfica Inscripciones (Historial por fecha de inscripción)
        cursor.execute("""
            SELECT FORMAT(fecha_inscripcion, 'yyyy-MM'), COUNT(inscripcion_id)
            FROM Inscripciones
            GROUP BY FORMAT(fecha_inscripcion, 'yyyy-MM')
            ORDER BY FORMAT(fecha_inscripcion, 'yyyy-MM')
        """)
        for rin in cursor.fetchall():
            inscripciones_labels.append(rin[0] or 'Sin Fecha')
            inscripciones_data.append(rin[1])

        # 5. Gráfica Evaluaciones (Aprobados >= 70 vs Reprobados < 70)
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN calificacion >= 70 THEN 1 ELSE 0 END),
                SUM(CASE WHEN calificacion < 70 THEN 1 ELSE 0 END)
            FROM Evaluaciones
        """)
        row_ev = cursor.fetchone()
        if row_ev:
            eval_aprobadas = row_ev[0] or 0
            eval_reprobadas = row_ev[1] or 0

        # 6. Gráfica Pagos (Ingresos por fecha de pago)
        cursor.execute("""
            SELECT FORMAT(fecha_pago, 'yyyy-MM'), SUM(monto)
            FROM Pagos
            WHERE estado = 'PAGADO'
            GROUP BY FORMAT(fecha_pago, 'yyyy-MM')
            ORDER BY FORMAT(fecha_pago, 'yyyy-MM')
        """)
        for rp in cursor.fetchall():
            pagos_labels.append(rp[0] or 'Sin Fecha')
            pagos_data.append(float(rp[1]) if rp[1] else 0.0)

        # Lógica existente para tus tablas de reportes
        if tipo_reporte == 'estudiantes':
            columnas = ['ID', 'Nombre Completo', 'Documento', 'Correo', 'Teléfono', 'Direccion', 'Fecha Registro', 'Estado']
            cursor.execute("""
                SELECT estudiante_id, nombre_completo, numero_documento, email, telefono, direccion, fecha_registro, estado 
                FROM Estudiantes 
                WHERE fecha_registro BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
            
        elif tipo_reporte == 'cursos':
            columnas = ['ID Curso', 'Nombre', 'Categoría', 'Duración (Horas)', 'Costo', 'Estado']
            cursor.execute("""
                SELECT curso_id, nombre_curso, categoria, duracion_horas, costo, estado 
                FROM Cursos
            """)

        elif tipo_reporte == 'instructores':
            columnas = ['ID', 'Nombre', 'Especialidad', 'Cédula Profesional', 'Correo', 'Teléfono', 'Dirección', 'Estado']
            cursor.execute("""
                SELECT instructor_id, nombre_completo, especialidad, cedula_profesional, email, telefono, direccion, estado 
                FROM Instructores
            """)
            
        elif tipo_reporte == 'inscripciones':
            columnas = ['Inscripción', 'Estudiante', 'Curso', 'Instructor', 'Folio', 'Fecha', 'Estado', 'Total Pago']
            cursor.execute("""
                SELECT i.inscripcion_id, e.nombre_completo, c.nombre_curso, ins.nombre_completo, i.folio_inscripcion, i.fecha_inscripcion, i.estado, i.total_pago
                FROM Inscripciones i
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
                JOIN Cursos c ON i.curso_id = c.curso_id
                LEFT JOIN Instructores ins ON c.instructor_id = ins.instructor_id
                WHERE i.fecha_inscripcion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])

        elif tipo_reporte == 'evaluaciones':
            columnas = ['Evaluación ID', 'Estudiante', 'Curso', 'Calificación', 'Comentarios', 'Fecha Evaluación']
            cursor.execute("""
                SELECT ev.evaluacion_id, e.nombre_completo, c.nombre_curso, ev.calificacion, 
                       CAST(ev.comentarios AS VARBINARY(MAX)), ev.fecha_evaluacion 
                FROM Evaluaciones ev 
                JOIN Inscripciones i ON ev.inscripcion_id = i.inscripcion_id
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id 
                JOIN Cursos c ON i.curso_id = c.curso_id 
                WHERE ev.fecha_evaluacion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
            
        elif tipo_reporte == 'pagos':
            columnas = ['Pago ID', 'Estudiante', 'Curso', 'Fecha Pago', 'Monto', 'Referencia', 'Estado']
            cursor.execute("""
                SELECT p.pago_id, e.nombre_completo, c.nombre_curso, p.fecha_pago, p.monto, p.referencia_pago, p.estado
                FROM Pagos p
                JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
                JOIN Cursos c ON i.curso_id = c.curso_id
                WHERE p.fecha_pago BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])

        raw_resultados = cursor.fetchall()
        
        resultados = []
        for fila in raw_resultados:
            fila_lista = list(fila)
            # Tus funciones de cifrado se mantienen intactas
            resultados.append(tuple(fila_lista))

    context = {
        'tipo_reporte': tipo_reporte,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'columnas': columnas,
        'resultados': resultados,
        'nombres': ['Reporte de Estudiantes', 'Reporte de Cursos','Reporte de Instructores', 'Reporte de Inscripciones', 'Reporte de Evaluaciones', 'Reporte de Pagos'],
        # Variables enviadas al template para las gráficas automáticas
        'total_activos': total_activos,
        'total_inactivos': total_inactivos,
        'cursos_labels': cursos_labels,
        'cursos_data': cursos_data,
        'instructores_labels': instructores_labels,
        'instructores_data': instructores_data,
        'inscripciones_labels': inscripciones_labels,
        'inscripciones_data': inscripciones_data,
        'eval_aprobadas': eval_aprobadas,
        'eval_reprobadas': eval_reprobadas,
        'pagos_labels': pagos_labels,
        'pagos_data': pagos_data,
    }

    return render(request, 'inventario/reportes.html', context)
@login_required
def generar_reporte_pdf(request):
    tipo_reporte = request.GET.get('tipo', 'estudiantes')
    fecha_inicio = request.GET.get('fecha_inicio', '2026-01-01')
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    resultados = []
    columnas = []

    with connection.cursor() as cursor:
        if tipo_reporte == 'estudiantes':
            columnas = ['ID', 'Nombre Completo', 'Documento (Cifrado)', 'Correo', 'Teléfono', 'Dirección', 'Fecha Registro', 'Estado']
            cursor.execute("SELECT estudiante_id, nombre_completo, numero_documento, email, telefono, direccion, fecha_registro, estado FROM Estudiantes WHERE fecha_registro BETWEEN %s AND %s", [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'cursos':
            columnas = ['ID Curso', 'Nombre', 'Categoría', 'Duración (Horas)', 'Costo', 'Estado']
            cursor.execute("SELECT curso_id, nombre_curso, categoria, duracion_horas, costo, estado FROM Cursos")

        elif tipo_reporte == 'instructores':
            columnas = ['ID', 'Nombre', 'Especialidad', 'Cédula Prof. (Cifrada)', 'Correo', 'Teléfono', 'Dirección', 'Estado']
            cursor.execute("SELECT instructor_id, nombre_completo, especialidad, cedula_profesional, email, telefono, direccion, estado FROM Instructores")  

        elif tipo_reporte == 'inscripciones':
            columnas = ['Inscripción', 'Estudiante', 'Curso', 'Instructor', 'Folio', 'Fecha', 'Estado', 'Total Pago']
            cursor.execute("""
                SELECT i.inscripcion_id, e.nombre_completo, c.nombre_curso, ins.nombre_completo, i.folio_inscripcion, i.fecha_inscripcion, i.estado, i.total_pago 
                FROM Inscripciones i 
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id 
                JOIN Cursos c ON i.curso_id = c.curso_id 
                LEFT JOIN Instructores ins ON c.instructor_id = ins.instructor_id 
                WHERE i.fecha_inscripcion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'evaluaciones':
            columnas = ['Evaluación ID', 'Estudiante', 'Curso', 'Calificación', 'Comentarios (Cifrados)', 'Fecha Evaluación']
            cursor.execute("""
                SELECT ev.evaluacion_id, e.nombre_completo, c.nombre_curso, ev.calificacion, 
                       CAST(ev.comentarios AS VARBINARY(MAX)), ev.fecha_evaluacion 
                FROM Evaluaciones ev 
                JOIN Inscripciones i ON ev.inscripcion_id = i.inscripcion_id
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id 
                JOIN Cursos c ON i.curso_id = c.curso_id 
                WHERE ev.fecha_evaluacion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'pagos':
            columnas = ['Pago ID', 'Estudiante', 'Curso', 'Fecha Pago', 'Monto', 'Referencia', 'Estado']
            cursor.execute("SELECT p.pago_id, e.nombre_completo, c.nombre_curso, p.fecha_pago, p.monto, p.referencia_pago, p.estado FROM Pagos p JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id JOIN Cursos c ON i.curso_id = c.curso_id WHERE p.fecha_pago BETWEEN %s AND %s", [fecha_inicio, fecha_fin])

        raw_resultados = cursor.fetchall()
        
        resultados = []
        for fila in raw_resultados:
            fila_lista = list(fila)
            
            # Función auxiliar interna para sanitizar y compactar valores cifrados (bytes o strings largos)
            def compactar_cifrado(val):
                if val is None:
                    return ""
                if isinstance(val, bytes):
                    # Convertir bytes a Base64 legible y recortarlo para que quepa en la celda
                    encoded = base64.b64encode(val).decode('utf-8')
                else:
                    encoded = str(val)
                
                # Si excede una longitud razonable para una celda de tabla, recortarlo de forma segura
                if len(encoded) > 40:
                    return encoded[:37] + "..."
                return encoded

            if tipo_reporte == 'estudiantes':
                fila_lista[2] = compactar_cifrado(cifrar_valor(fila_lista[2]))
            elif tipo_reporte == 'instructores':
                fila_lista[3] = compactar_cifrado(cifrar_valor(fila_lista[3]))   
            elif tipo_reporte == 'evaluaciones':
                fila_lista[4] = compactar_cifrado(cifrar_valor(fila_lista[4]))
            elif tipo_reporte == 'pagos':
                fila_lista[5] = compactar_cifrado(cifrar_valor(fila_lista[5]))
                
            resultados.append(tuple(fila_lista))

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"reporte_admin_{tipo_reporte}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    file_path = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(letter),
        rightMargin=30, leftMargin=30,
        topMargin=80, bottomMargin=40
    )

    styles = getSampleStyleSheet()
    
    style_header = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1
    )
    
    style_cell = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        textColor=colors.HexColor("#1f2937")
    )

    style_cell_cifrado = ParagraphStyle(
        'CellCifradoStyle',
        parent=styles['Normal'],
        fontName='Courier',
        fontSize=6,
        textColor=colors.HexColor("#374151")
    )

    elementos = []
    header_row = [Paragraph(col, style_header) for col in columnas]
    data_rows = [header_row]

    for fila in resultados:
        if tipo_reporte == 'estudiantes':
            u_id, nombre, doc_cifrado, correo, tel, direccion, fecha, estado = fila
            fila_cells = [
                Paragraph(str(u_id) if u_id is not None else "", style_cell),
                Paragraph(str(nombre) if nombre is not None else "", style_cell),
                Paragraph(escape(str(doc_cifrado)), style_cell_cifrado),
                Paragraph(str(correo) if correo is not None else "", style_cell),
                Paragraph(str(tel) if tel is not None else "", style_cell),
                Paragraph(str(direccion) if direccion is not None else "", style_cell),
                Paragraph(str(fecha) if fecha is not None else "", style_cell),
                Paragraph(str(estado) if estado is not None else "", style_cell),
            ]
        elif tipo_reporte == 'instructores':
            u_id, nombre, especialidad, ced_cifrada, correo, tel, direccion, estado = fila
            fila_cells = [
                Paragraph(str(u_id) if u_id is not None else "", style_cell),
                Paragraph(str(nombre) if nombre is not None else "", style_cell),
                Paragraph(str(especialidad) if especialidad is not None else "", style_cell),
                Paragraph(escape(str(ced_cifrada)), style_cell_cifrado),
                Paragraph(str(correo) if correo is not None else "", style_cell),
                Paragraph(str(tel) if tel is not None else "", style_cell),
                Paragraph(str(direccion) if direccion is not None else "", style_cell),
                Paragraph(str(estado) if estado is not None else "", style_cell),
            ]
        else:
            fila_cells = []
            for i, val in enumerate(fila):
                val_str = str(val) if val is not None else ""
                es_cifrado = (
                    (tipo_reporte == 'inscripciones' and i == 4) or
                    (tipo_reporte == 'evaluaciones' and i == 4) or
                    (tipo_reporte == 'pagos' and i == 5)
                )
                estilo_usar = style_cell_cifrado if es_cifrado else style_cell
                fila_cells.append(Paragraph(escape(val_str), estilo_usar))
            
        data_rows.append(fila_cells)
        if tipo_reporte == 'estudiantes':
          col_widths = [30, 95, 115, 135, 110, 65, 100, 52]
        elif tipo_reporte == 'instructores':
          col_widths = [30, 95, 115, 135, 110, 65, 100, 52]
        elif tipo_reporte == 'evaluaciones':
          col_widths = [50, 110, 110, 50, 190, 80]  # Ajusta el ancho de la columna 4 (comentarios) a 190
        else:
           col_widths = None
    t = LongTable(data_rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#312e81")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
    ]))

    elementos.append(t)

    def header_footer(canvas_obj, document):
        canvas_obj.saveState()
        width_l, height_l = landscape(letter)
        
        canvas_obj.setFillColor(colors.HexColor("#312e81"))
        canvas_obj.rect(0, height_l - 60, width_l, 60, fill=1, stroke=0)

        canvas_obj.setFillColor(colors.white)
        canvas_obj.setFont("Helvetica-Bold", 14)
        canvas_obj.drawString(30, height_l - 25, f"SIPEL - REPORTE DE ADMIN ({tipo_reporte.upper()})")
        canvas_obj.setFont("Helvetica", 8.5)
        canvas_obj.drawString(30, height_l - 45, f"Rango de Fechas: {fecha_inicio} al {fecha_fin}")

        canvas_obj.setStrokeColor(colors.HexColor("#d1d5db"))
        canvas_obj.setLineWidth(1)
        canvas_obj.line(30, 25, width_l - 30, 25)
        
        canvas_obj.setFont("Helvetica-Oblique", 7.5)
        canvas_obj.setFillColor(colors.HexColor("#6b7280"))
        canvas_obj.drawString(30, 15, "Documento generado automáticamente por el sistema SIPEL.")
        canvas_obj.restoreState()

    doc.build(elementos, onFirstPage=header_footer, onLaterPages=header_footer)

    with open(file_path, 'rb') as pdf_file:
        pdf_data = pdf_file.read()

    return HttpResponse(
        pdf_data,
        content_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="{filename}"'}
    )
@login_required
def generar_reporte_excel(request):
    tipo_reporte = request.GET.get('tipo', 'estudiantes')
    fecha_inicio = request.GET.get('fecha_inicio', '2026-01-01')
    fecha_fin = request.GET.get('fecha_fin', datetime.now().strftime('%Y-%m-%d'))
    
    resultados = []
    columnas = []

    with connection.cursor() as cursor:
        if tipo_reporte == 'estudiantes':
            columnas = ['ID', 'Nombre Completo', 'Documento (Cifrado)', 'Correo', 'Teléfono', 'Dirección', 'Fecha Registro', 'Estado']
            cursor.execute("SELECT estudiante_id, nombre_completo, numero_documento, email, telefono, direccion, fecha_registro, estado FROM Estudiantes WHERE fecha_registro BETWEEN %s AND %s", [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'cursos':
            columnas = ['ID Curso', 'Nombre', 'Categoría', 'Duración (Horas)', 'Costo', 'Estado']
            cursor.execute("SELECT curso_id, nombre_curso, categoria, duracion_horas, costo, estado FROM Cursos")
        elif tipo_reporte == 'instructores':
            columnas = ['ID', 'Nombre', 'Especialidad', 'Cédula Prof. (Cifrada)', 'Correo', 'Teléfono', 'Dirección', 'Estado']
            cursor.execute("SELECT instructor_id, nombre_completo, especialidad, cedula_profesional, email, telefono, direccion, estado FROM Instructores")
        elif tipo_reporte == 'inscripciones':
            columnas = ['Inscripción', 'Estudiante', 'Curso', 'Instructor', 'Folio', 'Fecha', 'Estado', 'Total Pago']
            cursor.execute("""
                SELECT i.inscripcion_id, e.nombre_completo, c.nombre_curso, ins.nombre_completo, i.folio_inscripcion, i.fecha_inscripcion, i.estado, i.total_pago 
                FROM Inscripciones i 
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id 
                JOIN Cursos c ON i.curso_id = c.curso_id 
                LEFT JOIN Instructores ins ON c.instructor_id = ins.instructor_id 
                WHERE i.fecha_inscripcion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'evaluaciones':
            columnas = ['Evaluación ID', 'Estudiante', 'Curso', 'Calificación', 'Comentarios (Cifrados)', 'Fecha Evaluación']
            cursor.execute("""
                SELECT ev.evaluacion_id, e.nombre_completo, c.nombre_curso, ev.calificacion, 
                       CAST(ev.comentarios AS VARBINARY(MAX)), ev.fecha_evaluacion 
                FROM Evaluaciones ev 
                JOIN Inscripciones i ON ev.inscripcion_id = i.inscripcion_id
                JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id 
                JOIN Cursos c ON i.curso_id = c.curso_id 
                WHERE ev.fecha_evaluacion BETWEEN %s AND %s
            """, [fecha_inicio, fecha_fin])
        elif tipo_reporte == 'pagos':
            columnas = ['Pago ID', 'Estudiante', 'Curso', 'Fecha Pago', 'Monto', 'Referencia', 'Estado']
            cursor.execute("SELECT p.pago_id, e.nombre_completo, c.nombre_curso, p.fecha_pago, p.monto, p.referencia_pago, p.estado FROM Pagos p JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id JOIN Cursos c ON i.curso_id = c.curso_id WHERE p.fecha_pago BETWEEN %s AND %s", [fecha_inicio, fecha_fin])

        raw_resultados = cursor.fetchall()
        
        resultados = []
        for fila in raw_resultados:
            fila_lista = list(fila)
            
            # Función auxiliar interna idéntica a la del PDF para compactar valores binarios/cifrados
            def compactar_cifrado(val):
                if val is None:
                    return ""
                if isinstance(val, bytes):
                    encoded = base64.b64encode(val).decode('utf-8')
                else:
                    encoded = str(val)
                
                if len(encoded) > 40:
                    return encoded[:37] + "..."
                return encoded

            if tipo_reporte == 'estudiantes':
                fila_lista[2] = compactar_cifrado(cifrar_valor(fila_lista[2]))
            elif tipo_reporte == 'instructores':
                fila_lista[3] = compactar_cifrado(cifrar_valor(fila_lista[3]))
            elif tipo_reporte == 'evaluaciones':
                fila_lista[4] = compactar_cifrado(cifrar_valor(fila_lista[4]))
            elif tipo_reporte == 'pagos':
                fila_lista[5] = compactar_cifrado(cifrar_valor(fila_lista[5]))
                
            resultados.append(tuple(fila_lista))

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo_reporte.capitalize()

    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")
    header_font = Font(name="Helvetica", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="Helvetica", size=9, color="1F2937")
    cifrado_font = Font(name="Courier New", size=8, color="374151")
   
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
   
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    max_col_letter = get_column_letter(len(columnas))

    # Título principal adaptado al número de columnas
    ws.append([f"SIPEL - REPORTE DE ADMIN ({tipo_reporte.upper()})"])
    ws.merge_cells(f"A1:{max_col_letter}1")
    cell_title = ws["A1"]
    cell_title.font = Font(name="Helvetica", size=14, bold=True, color="FFFFFF")
    cell_title.fill = header_fill
    cell_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # Subtítulo de rango de fechas
    ws.append([f"Rango de Fechas: {fecha_inicio} al {fecha_fin}"])
    ws.merge_cells(f"A2:{max_col_letter}2")
    cell_sub = ws["A2"]
    cell_sub.font = Font(name="Helvetica", size=9, italic=True, color="374151")
    cell_sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    ws.append([])
    ws.row_dimensions[3].height = 10

    # Cabeceras de tabla
    ws.append(columnas)
    ws.row_dimensions[4].height = 25
    for col_num in range(1, len(columnas) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    row_idx = 5
    for fila in resultados:
        valores_str = [str(val) if val is not None else "" for val in fila]
        ws.append(valores_str)
        
        ws.row_dimensions[row_idx].height = 30
        
        for col_num in range(1, len(valores_str) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            
            # Identificar columna cifrada de forma general según el tipo de reporte
            es_cifrado = (
                (tipo_reporte == 'estudiantes' and col_num == 3) or
                (tipo_reporte == 'instructores' and col_num == 4) or
                (tipo_reporte == 'evaluaciones' and col_num == 5) or
                (tipo_reporte == 'pagos' and col_num == 6)
            )

            if es_cifrado:
                cell.font = cifrado_font
                cell.alignment = align_left
            elif col_num == 1:
                cell.font = cell_font
                cell.alignment = align_center
            else:
                cell.font = cell_font
                cell.alignment = align_left
                
        row_idx += 1

    # Anchos de columna personalizados
    if tipo_reporte in ['estudiantes', 'instructores']:
        ws.column_dimensions['A'].width = 8   # ID
        ws.column_dimensions['B'].width = 25  # Nombre Completo
        ws.column_dimensions['C'].width = 35  # Documento cifrado recortado
        ws.column_dimensions['D'].width = 25  # Correo
        ws.column_dimensions['E'].width = 15  # Teléfono
        ws.column_dimensions['F'].width = 30  # Dirección
        ws.column_dimensions['G'].width = 20  # Fecha Registro
        ws.column_dimensions['H'].width = 15  # Estado
    else:
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_length + 4, 15)

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"reporte_admin_{tipo_reporte}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(reports_dir, filename)

    wb.save(file_path)

    with open(file_path, 'rb') as excel_file:
        excel_data = excel_file.read()

    return HttpResponse(
        excel_data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )
#---------usuarios-----------------
def gestion_usuarios(request):
    usuarios = User.objects.all().order_by('id')

    # Filtro por texto (búsqueda de usuario o nombre)
    q = request.GET.get('q')
    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) | 
            Q(first_name__icontains=q) | 
            Q(last_name__icontains=q)
        )

    # Filtro por estado (activo/inactivo)
    estado = request.GET.get('estado')
    if estado == 'activo':
        usuarios = usuarios.filter(is_active=True)
    elif estado == 'inactivo':
        usuarios = usuarios.filter(is_active=False)

    # Capturar el rol seleccionado por los botones (Por defecto 'Administrador')
    rol_seleccionado = request.GET.get('rol', 'Administrador')

    # Filtrar la lista general según el botón activo
    usuarios_filtrados = usuarios.filter(groups__name=rol_seleccionado)

    contexto = {
        'usuarios': usuarios_filtrados,
        'rol_seleccionado': rol_seleccionado,
    }
    return render(request, 'admin_sistema/usuarios.html', contexto)
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    rol_actual = usuario.groups.first().name if usuario.groups.exists() else "Sin rol"

    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        nueva_contrasena = request.POST.get('password', '').strip()

        if form.is_valid():
            usuario_actualizado = form.save(commit=False)
            
            if nueva_contrasena:
                usuario_actualizado.set_password(nueva_contrasena)
            
            usuario_actualizado.save()
            
            # Sincronización automática de nombres de usuario
            if hasattr(usuario_actualizado, 'estudiante'):
                estudiante = usuario_actualizado.estudiante
                estudiante.usuario = usuario_actualizado.username
                estudiante.save()
            elif hasattr(usuario_actualizado, 'instructor'):
                instructor = usuario_actualizado.instructor
                instructor.usuario = usuario_actualizado.username
                instructor.save()

            messages.success(request, 'Credenciales de usuario actualizadas correctamente.')
            return redirect('inventario:gestion_usuarios')
    else:
        form = UsuarioForm(instance=usuario)
        
    context = {
        'form': form, 
        'usuario': usuario,
        'rol_actual': rol_actual
    }
    return render(request, 'inventario/editar_usuario.html', context)

def eliminar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado correctamente')
        return redirect('inventario:gestion_usuarios')
        
    return render(request, 'inventario/eliminar_usuario.html', {'usuario': usuario})

def custom_logout_view(request):
    logout(request)
    return redirect('core:index')
    
# --- VISTAS POR ROL ---
@login_required
@role_required(allowed_roles=['Estudiante'])
def registros_estudiante(request):
    usuario = request.user
    
    # Nombre para el saludo
    nombre_mostrar = usuario.first_name if usuario.first_name else usuario.username
    
    try:
        # Usamos usuario_auth=usuario que es el campo correcto de tu modelo Estudiante
        estudiante_obj = Estudiante.objects.get(usuario_auth=usuario)
        
        # Filtramos las evaluaciones a través de la inscripción del estudiante
        evaluaciones_estudiante = Evaluacion.objects.filter(inscripcion__estudiante=estudiante_obj)
        resultado = evaluaciones_estudiante.aggregate(promedio_general=Avg('calificacion'))
        promedio_actual = resultado['promedio_general']
    except Estudiante.DoesNotExist:
        promedio_actual = None
    
    # Si tiene calificaciones, redondeamos a 2 decimales; si no, dejamos 0.0
    if promedio_actual is not None:
        promedio_actual = round(promedio_actual, 2)
    else:
        promedio_actual = 0.0

    context = {
        'nombre_estudiante': nombre_mostrar,
        'promedio_actual': promedio_actual,
        'usuario': usuario.username,
        'permiso': 'Estudiante',
    }
    
    return render(request, 'inventario/registros_estudiante.html', context)


@login_required
@role_required(allowed_roles=['Instructor'])
def registros_instructor(request):
    # 1. Obtenemos el perfil del instructor logueado
    instructor = Instructor.objects.filter(usuario_auth=request.user).first()
    
    # 2. Filtramos las inscripciones únicamente para los cursos que pertenecen a este instructor
    # (Si en tu modelo Curso el campo se llama distinto, ajústalo ej: curso__profesor=instructor)
    if instructor:
        ultimas_inscripciones = Inscripcion.objects.filter(curso__instructor=instructor)[:5]
        mis_cursos = Curso.objects.filter(instructor=instructor)
    else:
        ultimas_inscripciones = Inscripcion.objects.none()
        mis_cursos = Curso.objects.none()
    
    total_cursos = mis_cursos.count()
    total_inscripciones = ultimas_inscripciones.count()
    
    context = {
        'nombre_real': instructor.nombre_completo if instructor else request.user.get_full_name() or request.user.username,
        'permiso': 'Instructor',
        'usuario': request.user.username,
        'ultimas_inscripciones': ultimas_inscripciones,
        'mis_cursos': mis_cursos, # <-- Pasamos los cursos al contexto
        'total_cursos': total_cursos,
        'total_inscripciones': total_inscripciones,
    }
    return render(request, 'inventario/registros_instructor.html', context)
# --- VISTA PRINCIPAL REGISTROS de ADMIN ---
@login_required
def home_registros(request):
    # Contar los registros reales de la base de datos
    total_estudiantes = Estudiante.objects.count()
    total_cursos = Curso.objects.count()
    total_instructores = Instructor.objects.count()
    total_inscripciones = Inscripcion.objects.count()

    # Obtener los últimos registros agregados usando las llaves primarias reales de SQL
    ultimas_inscripciones = Inscripcion.objects.order_by('-inscripcion_id')[:3]
    
    # Corregido: Usar 'curso_id' en lugar de 'id' ya que esa es la llave primaria de Curso
    ultimos_cursos = Curso.objects.order_by('-curso_id')[:3] 
    
    # Basado en la estructura de instructores, su llave primaria es 'instructor_id'
    ultimos_instructores = Instructor.objects.order_by('-instructor_id')[:3]

    context = {
        'total_estudiantes': total_estudiantes,
        'total_cursos': total_cursos,
        'total_instructores': total_instructores,
        'total_inscripciones': total_inscripciones,
        
        # Listas recientes para el HTML
        'ultimas_inscripciones': ultimas_inscripciones,
        'ultimos_cursos': ultimos_cursos,
        'ultimos_instructores': ultimos_instructores,
    }
    
    return render(request, 'inventario/registros.html', context)
# --- VISTAS PARA INSTRUCTOR ---
@login_required
@role_required(allowed_roles=['Instructor'])
def mis_cursos(request):
    cursos = Curso.objects.filter(instructor__usuario=request.user)
    return render(request, 'inventario/mis_cursos.html', {'mis_cursos': cursos})

@login_required
@role_required(allowed_roles=['Instructor'])
def lista_estudiantes(request):
    user = request.user
    
    try:
        # Filtramos estrictamente usando la relación exacta de tu modelo Instructor
        inscripciones = Inscripcion.objects.filter(
            curso__instructor__usuario_auth=user
        ).distinct()
    except Exception as e:
        # Si ocurre algún inconveniente, mantenemos la lista vacía por seguridad
        inscripciones = Inscripcion.objects.none()

    # Barra de búsqueda opcional
    busqueda = request.GET.get('busqueda', '').strip()
    if busqueda:
        inscripciones = inscripciones.filter(
            Q(estudiante__nombre_completo__icontains=busqueda) |
            Q(estudiante__nombre__icontains=busqueda) |
            Q(curso__nombre_curso__icontains=busqueda)
        ).distinct()
        
    return render(request, 'inventario/lista_estudiantes.html', {'inscripciones': inscripciones})
from django.db import connection
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from core.decorators import role_required
from inventario.models import Curso

@login_required
@role_required(allowed_roles=['Instructor'])
def evaluar_alumnos(request):
    cursos_instructor = Curso.objects.filter(instructor__usuario=request.user)
    curso_id = request.GET.get('curso')
    busqueda = request.GET.get('q', '').strip()
    
    inscripciones = []
    
    with connection.cursor() as cursor:
        # 1. Abrir Master Key y Llave Simétrica para permitir el descifrado
        cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
        cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
        
        # 2. Consulta SQL usando auth_user para el nombre (ajusta 'first_name' y 'last_name' o el campo que use tu auth_user)
        sql = """
            SELECT 
                i.inscripcion_id,
                CONCAT(au.first_name, ' ', au.last_name) AS nombre_estudiante,
                c.nombre_curso,
                ev.calificacion,
                CAST(DecryptByKey(ev.comentarios) AS VARCHAR(MAX)) AS comentarios,
                ev.fecha_evaluacion
            FROM Inscripciones i
            INNER JOIN Cursos c ON i.curso_id = c.curso_id
            INNER JOIN Instructores inst ON c.instructor_id = inst.instructor_id
            INNER JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
            INNER JOIN auth_user au ON e.usuario_id = au.id
            LEFT JOIN Evaluaciones ev ON i.inscripcion_id = ev.inscripcion_id
            WHERE inst.usuario_id = %s
        """
        params = [request.user.pk]
        
        if curso_id:
            sql += " AND i.curso_id = %s"
            params.append(curso_id)
        if busqueda:
            sql += " AND (au.first_name LIKE %s OR au.last_name LIKE %s)"
            params.append(f"%{busqueda}%")
            params.append(f"%{busqueda}%")
            
        cursor.execute(sql, params)
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            data = dict(zip(columns, row))
            # Si el nombre completo se arma de first_name y last_name o si prefieres usar el username en caso de estar vacío:
            nombre_completo = data['nombre_estudiante'].strip()
            if not nombre_completo:
                nombre_completo = "Estudiante"
                
            inscripciones.append({
                'pk': data['inscripcion_id'],
                'estudiante': type('obj', (object,), {'nombre_completo': nombre_completo}),
                'curso': type('obj', (object,), {'nombre_curso': data['nombre_curso']}),
                'calificacion': data['calificacion'],
                'comentarios': data['comentarios'],
                'fecha_evaluacion': data['fecha_evaluacion']
            })
            
        # 3. Cerrar la llave simétrica
        cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")

    return render(request, 'inventario/evaluar_instructor.html', {
        'inscripciones': inscripciones, 
        'cursos': cursos_instructor
    })

@login_required
@role_required(allowed_roles=['Instructor'])
def registrar_nota(request, inscripcion_id):
    inscripcion = get_object_or_404(Inscripcion, pk=inscripcion_id)
    if request.method == 'POST':
        try:
            calificacion = float(request.POST.get('calificacion', 0))
            comentarios = request.POST.get('comentarios', '')
            
            # Usamos una transacción con SQL nativo para asegurar el cifrado simétrico en la sesión
            with transaction.atomic():
                with connection.cursor() as cursor:
                    # 1. Abrir la llave maestra con la contraseña configurada
                    cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
                    
                    # 2. Abrir la llave simétrica usando el certificado correspondiente
                    cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
                    
                    # 3. Verificar si ya existe una evaluación previa para esta inscripción
                    cursor.execute("SELECT COUNT(*) FROM Evaluaciones WHERE inscripcion_id = %s", [inscripcion.pk])
                    existe = cursor.fetchone()[0]
                    
                    if existe > 0:
                        # Si ya existe, actualizamos cifrando el comentario
                        cursor.execute("""
                            UPDATE Evaluaciones 
                            SET calificacion = %s, 
                                comentarios = EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), %s)), 
                                fecha_evaluacion = GETDATE()
                            WHERE inscripcion_id = %s
                        """, [calificacion, comentarios, inscripcion.pk])
                    else:
                        # Si no existe, insertamos cifrando el comentario
                        cursor.execute("""
                            INSERT INTO Evaluaciones (inscripcion_id, calificacion, comentarios, fecha_evaluacion)
                            VALUES (%s, %s, EncryptByKey(Key_GUID('ClaveDatos'), CONVERT(VARBINARY(MAX), %s)), GETDATE())
                        """, [inscripcion.pk, calificacion, comentarios])
                    
                    # 4. Cerrar la llave simétrica por seguridad
                    cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")

            messages.success(request, 'Nota registrada y cifrada correctamente.')
        except (ValueError, TypeError, Exception) as e:
            messages.error(request, f'Error al registrar la nota: {e}')
        return redirect('inventario:evaluar_alumnos') 
    return render(request, 'inventario/registrar_nota.html', {'inscripcion': inscripcion})
    #-----vista para estudiantes----------------------------


@login_required
@role_required(allowed_roles=['Estudiante'])
def mis_materias(request):
    try:
        estudiante_actual = Estudiante.objects.filter(usuario_auth=request.user).first()
        if not estudiante_actual:
            estudiante_actual = Estudiante.objects.filter(usuario_id=request.user.id).first()
            
        if not estudiante_actual:
            inscripciones = []
        else:
            # Traemos las inscripciones del estudiante
            inscripciones = Inscripcion.objects.filter(estudiante_id=estudiante_actual.estudiante_id).select_related('curso')
            
    except Exception as e:
        print(f"Error al cargar materias: {e}")
        inscripciones = []

    return render(request, 'inventario/mis_materias.html', {'inscripciones': inscripciones})
@login_required
@role_required(allowed_roles=['Estudiante'])
def mis_inscripciones(request):
    try:
        estudiante_actual = Estudiante.objects.get(usuario_auth=request.user)
    except Estudiante.DoesNotExist:
        estudiante_actual = None

    filtro_estado = request.GET.get('estado', '').strip()
    filtro_busqueda = request.GET.get('busqueda', '').strip()

    inscripciones_list = []
    if estudiante_actual:
        with connection.cursor() as cursor:
            # Quitamos el DecryptByKey para que la referencia_pago regrese en formato crudo/bytes (igual que el documento del estudiante)
            query = """
                SELECT 
                    c.nombre_curso,
                    c.categoria,
                    c.duracion_horas,
                    c.costo,
                    p.metodo_pago,
                    p.referencia_pago,
                    i.fecha_inscripcion,
                    p.fecha_pago,
                    p.monto,
                    ISNULL(p.estado, 'Pendiente') AS estado
                FROM Inscripciones i
                INNER JOIN Cursos c ON i.curso_id = c.curso_id
                LEFT JOIN pagos p ON p.inscripcion_id = i.inscripcion_id
                WHERE i.estudiante_id = %s
            """
            params = [estudiante_actual.estudiante_id]

            # Filtro opcional por búsqueda de nombre del curso
            if filtro_busqueda:
                query += " AND c.nombre_curso LIKE %s"
                params.append(f"%{filtro_busqueda}%")

            # Filtro opcional por estado
            if filtro_estado:
                query += " AND ISNULL(p.estado, 'Pendiente') = %s"
                params.append(filtro_estado)

            cursor.execute(query, params)
            
            columns = [col[0].lower() for col in cursor.description]
            inscripciones_list = [dict(zip(columns, row)) for row in cursor.fetchall()]

    cursos_disponibles = Curso.objects.all()

    context = {
        'inscripciones': inscripciones_list,
        'cursos_disponibles': cursos_disponibles,
        'total_inscripciones': len(inscripciones_list),
    }
    return render(request, 'inventario/mis_inscripciones.html', context)
@login_required
@role_required(allowed_roles=['Estudiante'])
def inscribir_curso(request):
    if request.method == 'POST':
            curso_id = request.POST.get('curso_id')
            metodo_pago = request.POST.get('metodo_pago')
            referencia_pago = request.POST.get('referencia_pago', '').strip()
            
            accion = request.POST.get('accion')
            estado_destino = 'Cancelada' if accion == 'cancelar' else 'Pagado'
            
            estudiante_actual = Estudiante.objects.filter(usuario_auth=request.user).first()
            if not estudiante_actual:
                estudiante_actual = Estudiante.objects.filter(usuario_id=request.user.id).first()
                
            if not estudiante_actual:
                raise Exception("No se encontró un perfil de estudiante vinculado a tu cuenta de usuario.")

            curso_actual = get_object_or_404(Curso, pk=curso_id)
            instructor_id = getattr(curso_actual, 'instructor_id', None)
            
            import random
            folio_generado = random.randint(2000, 9999)

            estudiante_db_id = int(estudiante_actual.estudiante_id)
            curso_db_id = int(curso_actual.curso_id)
            instructor_db_id = int(instructor_id) if instructor_id else "NULL"
            folio_db = int(folio_generado)
            costo_db = float(curso_actual.costo)
            
            metodo_clean = metodo_pago.replace("'", "''") if metodo_pago else ""
            referencia_clean = referencia_pago.replace("'", "''") if referencia_pago else ""
            estado_clean = estado_destino.replace("'", "''")

            sql_ejecucion = f"""
                EXEC sp_RegistrarInscripcionConPago 
                    @estudiante_id = {estudiante_db_id}, 
                    @curso_id = {curso_db_id}, 
                    @instructor_id = {instructor_db_id}, 
                    @folio_inscripcion = {folio_db}, 
                    @total_pago = {costo_db}, 
                    @estado_inscripcion = '{estado_clean}', 
                    @metodo_pago = '{metodo_clean}', 
                    @referencia_pago = '{referencia_clean}';
            """

            with connection.cursor() as cursor:
                cursor.execute(sql_ejecucion)
            
            if estado_destino == 'Cancelada':
                messages.warning(request, 'La inscripción ha sido marcada como cancelada.')
            else:
                messages.success(request, 'Te has inscrito y registrado tu pago correctamente.')
            
            return redirect('inventario:mis_inscripciones')
    return redirect('inventario:mis_inscripciones')
@login_required
@role_required(allowed_roles=['Estudiante'])
def mis_evaluaciones(request):
    estudiante = Estudiante.objects.filter(usuario_auth=request.user).first()
    
    evaluaciones = []
    if estudiante:
        with connection.cursor() as cursor:
            # 1. Abrir la Master Key y la Llave Simétrica en la sesión
            cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
            cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
            
            # 2. Consultar las evaluaciones descifrando el campo comentarios con DecryptByKey
            cursor.execute("""
                SELECT 
                    e.evaluacion_id,
                    e.inscripcion_id,
                    e.calificacion,
                    CAST(DecryptByKey(e.comentarios) AS VARCHAR(MAX)) AS comentarios,
                    e.fecha_evaluacion
                FROM Evaluaciones e
                INNER JOIN Inscripciones i ON e.inscripcion_id = i.inscripcion_id
                WHERE i.estudiante_id = %s
            """, [estudiante.estudiante_id])
            
            columns = [col[0] for col in cursor.description]
            evaluaciones = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # 3. Cerrar la llave simétrica por seguridad
            cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")

    context = {
        'evaluaciones': evaluaciones,
    }
    
    return render(request, 'inventario/mis_evaluaciones.html', context)
@login_required
@role_required(allowed_roles=['Estudiante'])
def mis_pagos(request):
    pagos = []
    
    with connection.cursor() as cursor:
        # Abrir llaves de seguridad para descifrar la referencia de pago
        cursor.execute("OPEN MASTER KEY DECRYPTION BY PASSWORD = 'MiPassword123*';")
        cursor.execute("OPEN SYMMETRIC KEY ClaveDatos DECRYPTION BY CERTIFICATE CertificadoDatos;")
        
        sql = """
            SELECT 
                p.pago_id,
                c.nombre_curso,
                p.fecha_pago,
                p.monto,
                p.estado,
                CAST(DecryptByKey(p.referencia_pago) AS VARCHAR(MAX)) AS referencia_pago
            FROM Pagos p
            INNER JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id
            INNER JOIN Cursos c ON i.curso_id = c.curso_id
            INNER JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
            WHERE e.usuario_id = %s
        """
        cursor.execute(sql, [request.user.pk])
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        
        for row in rows:
            data = dict(zip(columns, row))
            
            # Procesamiento seguro de la referencia descifrada por SQL
            ref_raw = data['referencia_pago']
            referencia_limpia = "N/A"
            
            if ref_raw:
                try:
                    if isinstance(ref_raw, memoryview):
                        referencia_limpia = ref_raw.tobytes().decode('utf-8')
                    elif isinstance(ref_raw, bytes):
                        referencia_limpia = ref_raw.decode('utf-8')
                    else:
                        referencia_limpia = str(ref_raw)
                except Exception:
                    referencia_limpia = "Error al procesar referencia"

            # Limpiamos espacios en blanco sobrantes devueltos por SQL Server
            estado_limpio = data['estado'].strip() if data['estado'] else ''

            pagos.append({
                'pago_id': data['pago_id'],
                'curso': data['nombre_curso'],
                'fecha': data['fecha_pago'],
                'monto': data['monto'],
                'estado': estado_limpio,
                'referencia': referencia_limpia
            })
            
        cursor.execute("CLOSE SYMMETRIC KEY ClaveDatos;")

    return render(request, 'inventario/mis_pagos.html', {'pagos': pagos})
@login_required
@role_required(['Estudiante'])
def descargar_comprobante_pdf(request, pago_id):
    # 1. Obtener los datos del pago (¡Sin descifrar, para que traiga los bytes cifrados!)
    with connection.cursor() as cursor:
        try:
            # Ya no es estrictamente necesario abrir la Master Key ni Symmetric Key 
            # a menos que requieras consultar otros campos cifrados.
            cursor.execute("""
                SELECT 
                    p.pago_id,
                    c.nombre_curso,
                    p.fecha_pago,
                    p.monto,
                    p.estado,
                    p.referencia_pago
                FROM Pagos p
                INNER JOIN Inscripciones i ON p.inscripcion_id = i.inscripcion_id
                INNER JOIN Cursos c ON i.curso_id = c.curso_id
                INNER JOIN Estudiantes e ON i.estudiante_id = e.estudiante_id
                WHERE e.usuario_id = %s AND p.pago_id = %s
            """, [request.user.id, pago_id])
            row = cursor.fetchone()
        finally:
            pass

    if not row:
        raise Http404("El comprobante no existe o no tienes permisos para verlo.")

    # Procesar los bytes cifrados de la referencia tal como se hizo con la cédula
    ref_raw = row[5]
    referencia_cifrada = "N/A"
    
    if ref_raw:
        try:
            if isinstance(ref_raw, memoryview):
                referencia_cifrada = str(ref_raw.tobytes())
            elif isinstance(ref_raw, bytes):
                referencia_cifrada = str(ref_raw)
            else:
                referencia_cifrada = str(ref_raw)
        except Exception:
            referencia_cifrada = "Error al procesar referencia"

    # Mapear los datos de la consulta incluyendo la referencia cifrada
    pago_data = {
        'pago_id': row[0],
        'curso': row[1],
        'fecha': row[2],
        'monto': row[3],
        'estado': row[4],
        'referencia': referencia_cifrada,
    }

    # 2. Definir la ruta de la carpeta media\reports
    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)  # Crear la carpeta si no existe

    # Nombre del archivo PDF basado en el ID del pago
    filename = f"comprobante_pago_estudiante_{pago_data['pago_id']}.pdf"
    file_path = os.path.join(reports_dir, filename)

    # 3. Generar el PDF utilizando ReportLab
    c = canvas.Canvas(file_path, pagesize=letter)
    width, height = letter

    # Encabezado estético
    c.setFillColor(colors.HexColor("#312e81"))  # Tono Índigo oscuro
    c.rect(0, height - 100, width, 100, fill=1, stroke=0)

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(50, height - 45, "SIPEL - COMPROBANTE DE PAGO")
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 70, f"Comprobante Oficial N°: 0000{pago_data['pago_id']}")

    # Cuerpo / Información del pago
    c.setFillColor(colors.HexColor("#1f2937"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, height - 150, "Detalles de la Transacción:")

    c.setFont("Helvetica", 12)
    y_position = height - 190
    space = 30

    detalles = [
        ("Curso Adquirido:", pago_data['curso']),
        ("Fecha y Hora:", pago_data['fecha'].strftime("%d/%m/%Y %H:%M") if pago_data['fecha'] else "N/A"),
        ("Monto Cancelado:", f"${pago_data['monto']}"),
        ("Estado del Pago:", pago_data['estado'].upper() if pago_data['estado'] else "N/A"),
        ("Referencia de Pago:", pago_data['referencia'])
    ]

    for label, value in detalles:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y_position, label)
        
        # Ajustamos el tamaño de fuente o usamos wrap si la cadena cifrada es muy larga en el PDF
        c.setFont("Helvetica", 9) # Fuente un poco más pequeña para que quepa el hash/cifrado
        c.drawString(200, y_position, str(value))
        y_position -= space

    # Línea divisoria inferior
    c.setStrokeColor(colors.HexColor("#d1d5db"))
    c.setLineWidth(1)
    c.line(50, 150, width - 50, 150)

    # Pie de página
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(colors.HexColor("#6b7280"))
    c.drawString(50, 120, "Este documento es un comprobante digital generado automáticamente por el sistema SIPEL.")
    c.drawString(50, 105, "Conserve este archivo para cualquier aclaración futura.")

    c.showPage()
    c.save()

    # 4. Servir el archivo generado para que se descargue / abra en el navegador
    return FileResponse(open(file_path, 'rb'), as_attachment=False, filename=filename)